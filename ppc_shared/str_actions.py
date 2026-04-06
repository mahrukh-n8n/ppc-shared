"""
STR actions — format view outputs for different consumers.

Functions to convert view data into:
  - optimization-ready outputs (action lists with recommendations)
  - audit-ready outputs (summary tables with severity flags)
  - app-ready JSON (serializable, camelCase keys)
  - Excel sheets (multi-tab workbook)
"""

import json
from datetime import datetime


def _to_camel_case(snake_str: str) -> str:
    """Convert snake_case to camelCase."""
    components = snake_str.split("_")
    return components[0] + "".join(x.title() for x in components[1:])


def _dict_to_camel(d: dict) -> dict:
    """Recursively convert dict keys from snake_case to camelCase."""
    if isinstance(d, dict):
        return {_to_camel_case(k): _dict_to_camel(v) for k, v in d.items()}
    if isinstance(d, list):
        return [_dict_to_camel(i) if isinstance(i, dict) else i for i in d]
    return d


def to_optimization_output(view_data: dict) -> dict:
    """Format view data for the optimization pipeline.

    Returns action-oriented output matching the format expected by
    downstream optimization prompts (04-action, 05-report).
    """
    output = {
        "generated_at": datetime.now().isoformat(),
        "promote": [],
        "negate": [],
        "bid_reduce": [],
        "cannibalization": [],
        "duplicates": [],
        "leakage_flags": [],
        "summary": {},
    }

    if "promote_candidates" in view_data:
        for p in view_data["promote_candidates"]:
            output["promote"].append(
                {
                    "term": p.get("customer_search_term", ""),
                    "orders": p.get("orders", 0),
                    "spend": p.get("spend", 0),
                    "sales": p.get("sales", 0),
                    "acos": p.get("acos"),
                    "cvr": p.get("cvr", 0),
                    "campaigns": p.get("campaigns", []),
                    "recommendation": p.get("action", "Add as Exact match keyword"),
                }
            )

    if "negate_candidates" in view_data:
        for n in view_data["negate_candidates"]:
            if n.get("all_already_negated"):
                continue
            output["negate"].append(
                {
                    "term": n.get("customer_search_term", ""),
                    "spend": n.get("spend", 0),
                    "clicks": n.get("clicks", 0),
                    "campaigns": n.get("needs_negative_in", n.get("campaigns", [])),
                    "match_type": "negative_exact"
                    if n.get("negate_type") == "Negative Exact"
                    else "negative_phrase",
                    "reason": f"Zero orders, ${n.get('spend', 0):.2f} spend",
                }
            )

    if "high_acos_converting" in view_data:
        for h in view_data["high_acos_converting"]:
            output["bid_reduce"].append(
                {
                    "term": h.get("customer_search_term", ""),
                    "acos": h.get("acos"),
                    "target_cpc": h.get("target_cpc"),
                    "campaigns": h.get("campaigns", []),
                    "action": h.get("action", "Reduce bid"),
                }
            )

    if "cannibalization" in view_data:
        output["cannibalization"] = view_data["cannibalization"]

    if "duplicate_terms" in view_data:
        output["duplicates"] = view_data["duplicate_terms"]

    if "leakage" in view_data:
        output["leakage_flags"] = [
            {
                "campaign": l["campaign_name"],
                "waste_ratio": l["waste_ratio"],
                "flags": l["flags"],
            }
            for l in view_data["leakage"]
            if l["flags"]
        ]

    output["summary"] = {
        "promote_count": len(output["promote"]),
        "negate_count": len(output["negate"]),
        "bid_reduce_count": len(output["bid_reduce"]),
        "cannibalization_count": len(output["cannibalization"]),
        "duplicate_count": len(output["duplicates"]),
        "leakage_flag_count": len(output["leakage_flags"]),
    }

    return output


def to_audit_output(view_data: dict) -> dict:
    """Format view data for audit reports — summary tables with severity flags."""
    output = {
        "generated_at": datetime.now().isoformat(),
        "sections": {},
    }

    # Funnel summary
    if "aggregated_terms" in view_data:
        terms = view_data["aggregated_terms"]
        zero = [t for t in terms if t.get("orders", 0) == 0]
        one = [t for t in terms if t.get("orders", 0) == 1]
        multi = [t for t in terms if t.get("orders", 0) >= 2]
        total_spend = sum(t.get("spend", 0) for t in terms)
        zero_spend = sum(t.get("spend", 0) for t in zero)

        output["sections"]["funnel"] = {
            "total_terms": len(terms),
            "zero_order": {"count": len(zero), "spend": round(zero_spend, 2)},
            "one_order": {
                "count": len(one),
                "spend": round(sum(t.get("spend", 0) for t in one), 2),
            },
            "multi_order": {
                "count": len(multi),
                "spend": round(sum(t.get("spend", 0) for t in multi), 2),
            },
            "waste_pct": round(zero_spend / total_spend * 100, 1)
            if total_spend > 0
            else 0,
            "waste_level": "Healthy"
            if zero_spend / total_spend * 100 < 20
            else (
                "Moderate"
                if zero_spend / total_spend * 100 < 35
                else ("High" if zero_spend / total_spend * 100 < 50 else "Critical")
            ),
        }

    # Action lists with severity
    if "promote_candidates" in view_data:
        output["sections"]["promote"] = {
            "count": len(view_data["promote_candidates"]),
            "items": view_data["promote_candidates"][:20],
        }

    if "negate_candidates" in view_data:
        net_new = [
            n
            for n in view_data["negate_candidates"]
            if not n.get("all_already_negated")
        ]
        output["sections"]["negate"] = {
            "count": len(net_new),
            "total_spend": round(sum(n.get("spend", 0) for n in net_new), 2),
            "items": net_new[:20],
        }

    if "cannibalization" in view_data:
        output["sections"]["cannibalization"] = {
            "count": len(view_data["cannibalization"]),
            "total_waste_spend": round(
                sum(c.get("waste_spend", 0) for c in view_data["cannibalization"]), 2
            ),
            "items": view_data["cannibalization"][:20],
        }

    if "leakage" in view_data:
        flagged = [l for l in view_data["leakage"] if l.get("flags")]
        output["sections"]["leakage"] = {
            "flagged_campaigns": len(flagged),
            "items": flagged,
        }

    return output


def to_app_json(view_data: dict) -> dict:
    """Format view data for the app — serializable, camelCase keys."""
    return _dict_to_camel(view_data)


def to_excel_sheets(view_data_dict: dict, output_path: str) -> str:
    """Write view data to a multi-tab Excel workbook.

    Args:
        view_data_dict: {sheet_name: [list_of_row_dicts]}
        output_path: Path to write the .xlsx file

    Returns:
        output_path
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Default styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for sheet_name, rows in view_data_dict.items():
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit

        if not rows:
            ws.append(["No data"])
            continue

        # Write headers
        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        # Write data
        for row_data in rows:
            values = []
            for h in headers:
                v = row_data.get(h)
                if isinstance(v, (list, dict)):
                    v = json.dumps(v)
                values.append(v)
            ws.append(values)

        # Apply borders to data rows
        for row_idx in range(2, len(rows) + 2):
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).border = thin_border

        # Auto-width (approximate)
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, min(len(rows) + 2, 102)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, min(len(str(cell_val)), 50))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (
                max_len + 2
            )

    wb.save(output_path)
    return output_path
